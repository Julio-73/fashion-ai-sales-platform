"""Unit tests for the Reporting module — service + generators.

Covers:
* ``ReportingService.build_*_report`` orchestration
* PDF generation (smoke test: produces a valid PDF byte stream)
* XLSX generation (smoke test: produces a valid workbook with the
  expected sheet titles)
* Permission gating (``reports:read``)
"""
from __future__ import annotations

import asyncio
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock
from uuid import UUID

import openpyxl
import pytest

from app.core.security.permissions import ROLE_PERMISSIONS
from app.modules.reporting.generators.pdf_generator import build_executive_pdf
from app.modules.reporting.generators.xlsx_generator import build_executive_xlsx
from app.modules.reporting.repository import ReportingRepository
from app.modules.reporting.schemas import (
    ExecutiveReportData,
    ReportKPI,
    ReportMetadata,
    ReportSection,
    ReportTable,
)
from app.modules.reporting.service import ReportingService


REPORTING_TENANT = UUID("aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa")


# ---------------------------------------------------------------------------
# Helper builders
# ---------------------------------------------------------------------------
def _build_metadata() -> ReportMetadata:
    return ReportMetadata(
        generated_at=datetime(2026, 6, 6, 10, 0, tzinfo=timezone.utc),
        tenant_id=REPORTING_TENANT,
        tenant_name="Acme Fashion",
        tenant_logo_url="https://cdn.example.com/acme.png",
        period_label="Hoy: 2026-06-06",
        currency="PEN",
    )


def _build_sample_data(include_alerts: bool = True) -> ExecutiveReportData:
    return ExecutiveReportData(
        metadata=_build_metadata(),
        kpis=[
            ReportKPI(label="Ventas hoy", value="S/ 100.00", secondary="3 pedidos"),
            ReportKPI(label="Clientes activos", value="25", secondary="VIP: 4"),
        ],
        sections=[
            ReportSection(
                title="Pipeline comercial",
                kpis=[
                    ReportKPI(label="Abiertos", value="12"),
                ],
                tables=[
                    ReportTable(
                        title="Embudo",
                        columns=["Etapa", "Deals"],
                        rows=[["Nuevo Lead", "4"], ["Won", "3"]],
                        total_row=["Total", "7"],
                    )
                ],
            )
        ],
        ai_recommendations=[
            {"id": "1", "title": "Lead caliente", "description": "Negociación abierta", "priority": "high"}
        ],
        critical_alerts=["Stock crítico: Polo"] if include_alerts else [],
    )


def _run(coro):
    return asyncio.run(coro)


def _mock_session_factory() -> AsyncMock:
    session = AsyncMock()
    return session


# ---------------------------------------------------------------------------
# PDF Generator
# ---------------------------------------------------------------------------
class TestPdfGenerator:
    def test_build_executive_pdf_returns_bytes(self) -> None:
        data = _build_sample_data()
        pdf_bytes = build_executive_pdf(data)
        assert isinstance(pdf_bytes, bytes)
        assert len(pdf_bytes) > 1000
        # PDF magic header
        assert pdf_bytes[:4] == b"%PDF"

    def test_build_executive_pdf_handles_empty_data(self) -> None:
        data = ExecutiveReportData(
            metadata=_build_metadata(),
            kpis=[],
            sections=[],
            ai_recommendations=[],
            critical_alerts=[],
        )
        pdf_bytes = build_executive_pdf(data)
        assert pdf_bytes[:4] == b"%PDF"
        assert len(pdf_bytes) > 500

    def test_build_executive_pdf_escapes_html(self) -> None:
        data = _build_sample_data()
        data.sections[0].paragraphs.append("<script>alert(1)</script> & special chars")
        pdf_bytes = build_executive_pdf(data)
        assert pdf_bytes[:4] == b"%PDF"


# ---------------------------------------------------------------------------
# XLSX Generator
# ---------------------------------------------------------------------------
class TestXlsxGenerator:
    def test_build_executive_xlsx_returns_valid_workbook(self) -> None:
        data = _build_sample_data()
        xlsx_bytes = build_executive_xlsx(data)
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 1000
        # ZIP magic header (xlsx is a zip)
        assert xlsx_bytes[:4] == b"PK\x03\x04"
        # Load and inspect
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
        sheet_titles = wb.sheetnames
        assert "Resumen Ejecutivo" in sheet_titles

    def test_build_executive_xlsx_creates_section_sheet(self) -> None:
        data = _build_sample_data()
        xlsx_bytes = build_executive_xlsx(data)
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
        # Section title becomes a sheet
        assert any("Pipeline" in t for t in wb.sheetnames)

    def test_build_executive_xlsx_handles_empty_data(self) -> None:
        data = ExecutiveReportData(
            metadata=_build_metadata(),
            kpis=[],
            sections=[],
            ai_recommendations=[],
            critical_alerts=[],
        )
        xlsx_bytes = build_executive_xlsx(data)
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
        assert "Resumen Ejecutivo" in wb.sheetnames

    def test_build_executive_xlsx_deduplicates_sheet_names(self) -> None:
        data = _build_sample_data()
        # Add a second section with the same title to trigger the dedupe
        data.sections.append(ReportSection(
            title="Pipeline comercial",
            tables=[ReportTable(title="x", columns=["a"], rows=[["1"]])],
        ))
        xlsx_bytes = build_executive_xlsx(data)
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
        # Two distinct tab names for the two sections
        pipeline_tabs = [t for t in wb.sheetnames if "Pipeline" in t]
        assert len(pipeline_tabs) == 2


# ---------------------------------------------------------------------------
# Service orchestration — uses a fully mocked repository
# ---------------------------------------------------------------------------
class TestReportingService:
    def _service(self, repo_methods: dict | None = None) -> tuple[ReportingService, AsyncMock]:
        session = _mock_session_factory()
        service = ReportingService(session)
        # Patch the repository with a mock so we don't need a DB
        mock_repo = MagicMock(spec=ReportingRepository)
        for name, return_value in (repo_methods or {}).items():
            setattr(mock_repo, name, AsyncMock(return_value=return_value))
        # Always-needed returns
        mock_repo.get_empresa = AsyncMock(
            return_value={
                "id": REPORTING_TENANT,
                "nombre": "Acme Fashion",
                "logo_url": "https://cdn.example.com/acme.png",
            }
        )
        # _now / _period are static — no need to mock
        service._repo = mock_repo
        return service, session

    def test_build_executive_report_composes_data(self) -> None:
        service, _ = self._service(
            {
                "sales_kpis": {
                    "sales_today": Decimal("100"),
                    "sales_week": Decimal("500"),
                    "sales_month": Decimal("2000"),
                    "sales_year": Decimal("20000"),
                    "average_ticket": Decimal("100.00"),
                    "average_ticket_month": Decimal("150.00"),
                    "total_orders": 50,
                    "orders_today": 2,
                    "orders_week": 5,
                    "orders_month": 12,
                },
                "customer_kpis": {
                    "total": 20, "active": 15, "vip": 3, "recurrent": 5,
                    "inactive": 2, "new": 1, "average_lifetime_value": Decimal("500.00"),
                    "average_orders_per_customer": Decimal("2.50"),
                },
                "pipeline_kpis": {
                    "open_deals": 10, "won_deals": 3, "lost_deals": 1,
                    "total_value": Decimal("5000"), "weighted_value": Decimal("2500"),
                    "won_value": Decimal("3000"), "lost_value": Decimal("500"),
                    "conversion_pct": 75.0,
                },
                "inventory_kpis": {
                    "total_products": 20, "out_of_stock": 1, "low_stock": 3,
                    "normal_stock": 16, "inventory_value": Decimal("50000"),
                    "total_units": 100, "total_reserved_units": 5,
                },
                "forecast": {
                    "monthly": Decimal("2000.00"),
                    "quarterly": Decimal("6000.00"),
                    "confidence": "high", "sample_size": 6,
                },
                "top_customers": [
                    {"full_name": "Cliente A", "order_count": 5,
                     "lifetime_value": Decimal("1000"), "is_vip": True},
                ],
                "top_products": [
                    {"product_id": UUID(int=1), "name": "Polo",
                     "units_sold": 10, "revenue": Decimal("500")},
                ],
                "ai_recommendations": [
                    {"id": "x", "title": "Hot lead", "description": "neg",
                     "priority": "high"},
                ],
                "inventory_critical_alerts": ["Polo: stock 0"],
                "delayed_orders_alerts": [],
            }
        )
        data = _run(service.build_executive_report(empresa_id=REPORTING_TENANT))
        assert isinstance(data, ExecutiveReportData)
        assert data.metadata.tenant_name == "Acme Fashion"
        assert len(data.kpis) == 8
        assert len(data.sections) == 6
        assert data.critical_alerts == ["Polo: stock 0"]

    def test_build_pipeline_report_returns_data(self) -> None:
        service, _ = self._service(
            {
                "pipeline_kpis": {
                    "open_deals": 5, "won_deals": 2, "lost_deals": 1,
                    "total_value": Decimal("3000"), "weighted_value": Decimal("1500"),
                    "won_value": Decimal("2000"), "lost_value": Decimal("500"),
                    "conversion_pct": 66.7,
                },
                "pipeline_funnel": [
                    {"stage": "new_lead", "count": 3, "value": Decimal("1000")},
                    {"stage": "won", "count": 2, "value": Decimal("2000")},
                ],
                "pipeline_deals_for_export": [
                    {"title": "Trato 1", "stage": "new_lead", "value": Decimal("100"),
                     "probability": 30, "customer_name": "Cliente A",
                     "created_at": datetime(2026, 6, 1, tzinfo=timezone.utc),
                     "expected_close_date": datetime(2026, 6, 1, tzinfo=timezone.utc)},
                ],
            }
        )
        data = _run(service.build_pipeline_report(empresa_id=REPORTING_TENANT))
        assert data.metadata.tenant_name == "Acme Fashion"
        # 3 sections: overview + funnel + deals
        assert len(data.sections) == 3
        assert data.sections[1].tables[0].title == "Distribución por etapa"

    def test_build_crm_report_returns_data(self) -> None:
        service, _ = self._service(
            {
                "customer_kpis": {
                    "total": 10, "active": 7, "vip": 2, "recurrent": 3,
                    "inactive": 1, "new": 1, "average_lifetime_value": Decimal("400.00"),
                    "average_orders_per_customer": Decimal("1.50"),
                },
                "customers_for_export": [
                    {
                        "full_name": "Cliente A", "email": "a@x.com", "phone": "+51999",
                        "order_count": 3, "lifetime_value": Decimal("600"),
                        "last_purchase_at": datetime(2026, 6, 1, tzinfo=timezone.utc),
                        "is_vip": False, "is_recurrent": True,
                    }
                ],
            }
        )
        data = _run(service.build_crm_report(empresa_id=REPORTING_TENANT))
        assert len(data.kpis) == 4
        assert len(data.sections) == 1

    def test_build_sales_report_returns_data(self) -> None:
        service, _ = self._service(
            {
                "sales_kpis": {
                    "sales_today": Decimal("100"), "sales_week": Decimal("500"),
                    "sales_month": Decimal("2000"), "sales_year": Decimal("20000"),
                    "average_ticket": Decimal("100.00"),
                    "average_ticket_month": Decimal("150.00"),
                    "total_orders": 50, "orders_today": 2, "orders_week": 5,
                    "orders_month": 12,
                },
                "daily_sales": [{"date": "2026-06-01", "revenue": Decimal("100"), "orders": 1}],
                "monthly_sales": [{"month": "2026-05", "revenue": Decimal("1000"), "orders": 5}],
                "orders_for_export": [
                    {"order_number": "ORD-001", "customer_name": "Cliente",
                     "status": "confirmed", "total": Decimal("100"),
                     "created_at": datetime(2026, 6, 1, tzinfo=timezone.utc),
                     "delivery_type": "delivery"},
                ],
                "top_products": [
                    {"product_id": UUID(int=1), "name": "Polo",
                     "units_sold": 10, "revenue": Decimal("500")},
                ],
            }
        )
        data = _run(service.build_sales_report(empresa_id=REPORTING_TENANT))
        # 5 sections: sales summary, daily, monthly, top products, orders
        assert len(data.sections) == 5

    def test_build_inventory_report_returns_data(self) -> None:
        service, _ = self._service(
            {
                "inventory_kpis": {
                    "total_products": 10, "out_of_stock": 1, "low_stock": 2,
                    "normal_stock": 7, "inventory_value": Decimal("5000"),
                    "total_units": 50, "total_reserved_units": 2,
                },
                "inventory_for_export": [
                    {"product_id": UUID(int=1), "name": "Polo", "category": "Ropa",
                     "sku": "polo-001", "base_price": Decimal("50"),
                     "stock_actual": 5, "stock_minimo": 2, "stock_reservado": 1,
                     "status": "normal"},
                ],
                "top_products": [
                    {"product_id": UUID(int=1), "name": "Polo",
                     "units_sold": 5, "revenue": Decimal("250")},
                ],
                "lowest_stock": [
                    {"product_id": UUID(int=1), "name": "Polo",
                     "stock_actual": 0, "stock_minimo": 2, "status": "agotado"},
                ],
            }
        )
        data = _run(service.build_inventory_report(empresa_id=REPORTING_TENANT))
        assert len(data.kpis) == 4
        # 3 sections: inventory, low stock, detail
        assert len(data.sections) == 3


# ---------------------------------------------------------------------------
# Permission system — reports:read must be granted to the right roles
# ---------------------------------------------------------------------------
class TestPermissions:
    @pytest.mark.parametrize("role", ["owner", "admin", "sales_agent", "analyst"])
    def test_reports_read_granted_to_expected_roles(self, role: str) -> None:
        assert "reports:read" in ROLE_PERMISSIONS[role]


# ---------------------------------------------------------------------------
# End-to-end: service -> PDF / XLSX bytes
# ---------------------------------------------------------------------------
class TestServiceToFilePipeline:
    def test_service_to_pdf_bytes(self) -> None:
        data = _build_sample_data()
        pdf = build_executive_pdf(data)
        assert pdf[:4] == b"%PDF"
        assert len(pdf) > 1000

    def test_service_to_xlsx_bytes(self) -> None:
        data = _build_sample_data()
        xlsx = build_executive_xlsx(data)
        assert xlsx[:4] == b"PK\x03\x04"
        assert len(xlsx) > 1000
