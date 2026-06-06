"""XLSX generator for the Reporting module.

Uses ``openpyxl`` to produce multi-sheet Excel workbooks. Each report
type renders a different combination of sheets, but the workbook
``__init__.py`` here exposes a single entry point —
``build_executive_xlsx(data) -> bytes`` — that the router uses for all
five report types.
"""
from __future__ import annotations

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.modules.reporting.schemas import (
    ExecutiveReportData,
    ReportSection,
    ReportTable,
)


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
_PRIMARY_HEX = "0F172A"
_ACCENT_HEX = "2563EB"
_HEADER_HEX = "1E293B"
_LIGHT_HEX = "F1F5F9"
_BORDER_HEX = "E2E8F0"
_SUCCESS_HEX = "10B981"
_WARNING_HEX = "F59E0B"
_DANGER_HEX = "EF4444"
_MUTED_HEX = "64748B"


def _header_font() -> Font:
    return Font(name="Calibri", size=11, bold=True, color="FFFFFF")


def _title_font() -> Font:
    return Font(name="Calibri", size=18, bold=True, color=_PRIMARY_HEX)


def _subtitle_font() -> Font:
    return Font(name="Calibri", size=10, italic=True, color=_MUTED_HEX)


def _kpi_label_font() -> Font:
    return Font(name="Calibri", size=9, color=_MUTED_HEX)


def _kpi_value_font() -> Font:
    return Font(name="Calibri", size=14, bold=True, color=_PRIMARY_HEX)


def _body_font() -> Font:
    return Font(name="Calibri", size=10, color=_PRIMARY_HEX)


def _total_font() -> Font:
    return Font(name="Calibri", size=10, bold=True, color=_PRIMARY_HEX)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_HEADER_HEX)


def _total_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_LIGHT_HEX)


def _kpi_fill() -> PatternFill:
    return PatternFill("solid", fgColor="DBEAFE")


def _thin_side() -> Side:
    return Side(style="thin", color=_BORDER_HEX)


def _all_border() -> Border:
    return Border(left=_thin_side(), right=_thin_side(), top=_thin_side(), bottom=_thin_side())


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def _write_cover_sheet(wb: Workbook, data: ExecutiveReportData) -> None:
    ws = wb.active
    ws.title = "Resumen Ejecutivo"
    ws["A1"] = f"Reporte Ejecutivo - {data.metadata.tenant_name}"
    ws["A1"].font = _title_font()
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = (
        f"Generado: {data.metadata.generated_at.strftime('%Y-%m-%d %H:%M UTC')}  -  "
        f"Periodo: {data.metadata.period_label}  -  Moneda: {data.metadata.currency}"
    )
    ws["A2"].font = _subtitle_font()
    ws.merge_cells("A2:D2")

    # KPI grid
    row = 4
    ws.cell(row=row, column=1, value="KPI").font = _header_font()
    ws.cell(row=row, column=2, value="Valor").font = _header_font()
    ws.cell(row=row, column=3, value="Detalle").font = _header_font()
    for col in range(1, 4):
        c = ws.cell(row=row, column=col)
        c.fill = _header_fill()
        c.alignment = _center()
        c.border = _all_border()
    row += 1
    for kpi in data.kpis:
        c1 = ws.cell(row=row, column=1, value=kpi.label)
        c1.font = _kpi_label_font()
        c1.fill = _kpi_fill()
        c1.alignment = _left()
        c1.border = _all_border()
        c2 = ws.cell(row=row, column=2, value=kpi.value)
        c2.font = _kpi_value_font()
        c2.alignment = _left()
        c2.border = _all_border()
        c3 = ws.cell(row=row, column=3, value=kpi.secondary or "")
        c3.font = _body_font()
        c3.alignment = _left()
        c3.border = _all_border()
        row += 1

    # Alertas
    if data.critical_alerts:
        row += 1
        ws.cell(row=row, column=1, value="Alertas críticas").font = _total_font()
        row += 1
        for a in data.critical_alerts:
            c = ws.cell(row=row, column=1, value=f"⚠ {a}")
            c.font = Font(name="Calibri", size=10, color=_DANGER_HEX)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1

    # AI recommendations
    if data.ai_recommendations:
        row += 1
        ws.cell(row=row, column=1, value="Recomendaciones IA").font = _total_font()
        row += 1
        ws.cell(row=row, column=1, value="Título").font = _header_font()
        ws.cell(row=row, column=2, value="Descripción").font = _header_font()
        ws.cell(row=row, column=3, value="Prioridad").font = _header_font()
        for col in range(1, 4):
            c = ws.cell(row=row, column=col)
            c.fill = _header_fill()
            c.alignment = _center()
            c.border = _all_border()
        row += 1
        for rec in data.ai_recommendations:
            ws.cell(row=row, column=1, value=rec.get("title", "")).font = _body_font()
            ws.cell(row=row, column=2, value=rec.get("description", "")).font = _body_font()
            prio = rec.get("priority", "medium")
            fill_color = (
                _DANGER_HEX
                if prio == "high"
                else _WARNING_HEX
                if prio == "medium"
                else _SUCCESS_HEX
            )
            cp = ws.cell(row=row, column=3, value=prio.upper())
            cp.font = Font(name="Calibri", size=10, bold=True, color=fill_color)
            cp.alignment = _center()
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = _all_border()
            row += 1

    _autosize(ws)


def _write_table_sheet(
    wb: Workbook,
    sheet_name: str,
    sections: Iterable[ReportSection],
) -> None:
    ws = wb.create_sheet(sheet_name[:31])
    row = 1
    for section in sections:
        if section.title:
            c = ws.cell(row=row, column=1, value=section.title)
            c.font = Font(name="Calibri", size=14, bold=True, color=_PRIMARY_HEX)
            row += 1
        for kpi in section.kpis:
            ws.cell(row=row, column=1, value=kpi.label).font = _kpi_label_font()
            ws.cell(row=row, column=2, value=kpi.value).font = _kpi_value_font()
            if kpi.secondary:
                ws.cell(
                    row=row, column=3, value=kpi.secondary
                ).font = _body_font()
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = _all_border()
            row += 1
        for table_block in section.tables:
            if table_block.title:
                ws.cell(row=row, column=1, value=table_block.title).font = _total_font()
                row += 1
            _write_data_block(ws, row, table_block)
            row += len(table_block.rows) + 1 + (1 if table_block.total_row else 0) + 1
        if section.alerts:
            for a in section.alerts:
                ws.cell(row=row, column=1, value=f"⚠ {a}").font = Font(
                    name="Calibri", size=10, color=_DANGER_HEX
                )
                row += 1
        row += 1
    _autosize(ws)


def _write_data_block(ws, start_row: int, table_block: ReportTable) -> None:
    for col_idx, header in enumerate(table_block.columns, start=1):
        c = ws.cell(row=start_row, column=col_idx, value=header)
        c.font = _header_font()
        c.fill = _header_fill()
        c.alignment = _center()
        c.border = _all_border()
    row = start_row + 1
    for r in table_block.rows:
        for col_idx, value in enumerate(r, start=1):
            c = ws.cell(row=row, column=col_idx, value=str(value))
            c.font = _body_font()
            c.alignment = _left()
            c.border = _all_border()
        row += 1
    if table_block.total_row:
        for col_idx, value in enumerate(table_block.total_row, start=1):
            c = ws.cell(row=row, column=col_idx, value=str(value))
            c.font = _total_font()
            c.fill = _total_fill()
            c.alignment = _left()
            c.border = _all_border()


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            length = max(len(line) for line in str(v).splitlines()) if str(v) else 0
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 50)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def build_executive_xlsx(data: ExecutiveReportData) -> bytes:
    """Build the multi-sheet XLSX report.

    Sheets written:
      1. ``Resumen Ejecutivo`` (cover + KPIs + alerts + AI recs)
      2. One sheet per section in ``data.sections``, with the section
         title used as the tab name (truncated to 31 chars).
    """
    wb = Workbook()
    _write_cover_sheet(wb, data)
    used_titles = {"Resumen Ejecutivo"}
    for section in data.sections:
        title = (section.title or "Sección").strip()[:31]
        # Excel requires unique sheet names; fall back to a counter.
        original = title
        counter = 2
        while title in used_titles:
            suffix = f" ({counter})"
            title = (original[: 31 - len(suffix)] + suffix)
            counter += 1
        used_titles.add(title)
        _write_table_sheet(wb, title, [section])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["build_executive_xlsx"]
